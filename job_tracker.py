import tkinter as tk
from tkinter import simpledialog, ttk
import openpyxl
from datetime import date
import os.path

# Check if the Excel file exists
if os.path.isfile('job_applications.xlsx'):
    # Load the existing workbook
    workbook = openpyxl.load_workbook('job_applications.xlsx')
else:
    # Create a new workbook
    workbook = openpyxl.Workbook()

# Get the active worksheet or create it if it doesn't exist
worksheet = workbook.active
if worksheet.title == "Sheet":
    worksheet.title = "Job Applications"

# Set column headers if they don't exist
if worksheet['A1'].value is None:
    worksheet['A1'] = 'Date'
    worksheet['B1'] = 'Company Name'
    worksheet['C1'] = 'Job Role'
    worksheet['D1'] = 'Contact Information'
    worksheet['E1'] = 'Job Status'
    worksheet['F1'] = 'Job Portal'
    worksheet['G1'] = 'Interview Date'

# Get the next available row
next_row = worksheet.max_row + 1

# Job status options
job_status_options = ["Save for Later", "Applied", "Interviewed", "Offer Received", "Rejected"]

def save_job():
    global next_row
    company_name = simpledialog.askstring("Company Name", "Enter the company name:")
    job_role = simpledialog.askstring("Job Role", "Enter the job role:")
    job_portal = simpledialog.askstring("Job Portal", "Enter the job portal or source:")
    contact_info = simpledialog.askstring("Contact Information", "Enter the contact information:")
    today = date.today()

    # Show dropdown to select job status
    job_status = tk.StringVar()
    job_status.set(job_status_options[0])  # Set the default option

    def select_job_status():
        selected_status = job_status_dropdown.get()
        job_status_window.destroy()

    job_status_window = tk.Toplevel(root)
    job_status_window.title("Select Job Status")
    job_status_dropdown = ttk.Combobox(job_status_window, textvariable=job_status, values=job_status_options)
    job_status_dropdown.pack(pady=10)
    select_button = tk.Button(job_status_window, text="Select", command=select_job_status)
    select_button.pack(pady=5)

    root.wait_window(job_status_window)  # Wait for the dropdown window to close

    # Write data to the worksheet
    worksheet.cell(row=next_row, column=1, value=company_name)
    worksheet.cell(row=next_row, column=2, value=job_role)
    worksheet.cell(row=next_row, column=3, value=today)
    worksheet.cell(row=next_row, column=4, value=contact_info)
    worksheet.cell(row=next_row, column=5, value=job_status.get())
    worksheet.cell(row=next_row, column=6, value=job_portal)

    # Save the workbook
    workbook.save('job_applications.xlsx')

    next_row += 1

# Create the main window
root = tk.Tk()
root.title("Job Application Tracker")
root.geometry("300x200")  # Set the window size

# Style the button
button_style = {
    "font": ("Helvetica", 14),
    "bg": "#4CAF50",
    "fg": "white",
    "padx": 20,
    "pady": 10,
    "relief": "raised",
    "borderwidth": 2
}

# Create the button
save_button = tk.Button(root, text="Save Job", command=save_job, **button_style)

# Add the button to the window
save_button.pack(pady=20)

# Run the main event loop
root.mainloop()